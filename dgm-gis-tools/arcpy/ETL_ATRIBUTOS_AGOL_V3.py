# -*- coding: utf-8 -*-
import arcpy
import os
import unicodedata
from datetime import datetime
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font

arcpy.env.overwriteOutput = True

# =========================================================
# CONFIG
# =========================================================
EXCEL_PATH = r"C:\CatastroMineroC\BaseDatosC\BD_RNM_AGOL.xlsx"
EXCEL_SHEET_CANDIDATES = ["Hoja1", "Hoja1$"]

TARGET_ATTR_NAME = "CM_ATRIBUTOS"
TARGET_BASE_NAME = "CM_BASE"

FIELD_EXPEDIENTE_EXCEL = "EXPEDIENTE"
FIELD_KEY = "EXPEDIENTE_KEY"

SYNC_FIELDS = [
    ("EXPEDIENTE_KEY", "TEXT", 100),
    ("Fecha_Actualizacion", "DATE", None),
    ("EN_REPORTE", "SHORT", None),
    ("FECHA_REPORTE", "DATE", None),
]

VALOR_SIN_DATOS = "SIN_DATOS"
VALOR_ARCHIVADO = "ARCHIVADO"

LOG_DIR = (
    r"C:\Users\echavarria\OneDrive - MINAE Costa Rica\2-REPORTES\Reportes_GIS\Rep_ETL"
)

# Prioridad para deduplicar target si existieran residuos anteriores
PREFERRED_DATE_FIELDS = ["FECHA_INGRESO", "Fecha_Actualizacion"]

# =========================================================
# RUTAS
# =========================================================
try:
    GDB_PATH = arcpy.mp.ArcGISProject("CURRENT").defaultGeodatabase
except Exception:
    GDB_PATH = arcpy.env.workspace

if not GDB_PATH or not arcpy.Exists(GDB_PATH):
    raise Exception(f"No se pudo determinar la GDB del proyecto. GDB_PATH={GDB_PATH}")

if not os.path.exists(LOG_DIR):
    os.makedirs(LOG_DIR, exist_ok=True)

TARGET_ATTR = os.path.join(GDB_PATH, TARGET_ATTR_NAME)
TARGET_BASE = os.path.join(GDB_PATH, TARGET_BASE_NAME)

RUN_TS = datetime.now().strftime("%Y%m%d_%H%M%S")
REPORT_XLSX = os.path.join(LOG_DIR, f"REPORTE_ETL_UNICO_{RUN_TS}.xlsx")

TEMP_GDB = arcpy.env.scratchGDB
STAGE_TABLE = os.path.join(TEMP_GDB, f"stg_etl_unico_{RUN_TS}")


# =========================================================
# UTILIDADES
# =========================================================
def normalize_text(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    s = " ".join(s.split())
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.upper()


def normalize_key(val) -> str:
    return normalize_text(val)


def exists(path) -> bool:
    return arcpy.Exists(path)


def field_names(table_path):
    return [f.name for f in arcpy.ListFields(table_path)]


def field_names_upper_map(table_path):
    return {normalize_text(f.name): f.name for f in arcpy.ListFields(table_path)}


def is_blocked_field(fname: str) -> bool:
    blocked = {"OBJECTID", "OID", "GLOBALID", "SHAPE", "SHAPE_LENGTH", "SHAPE_AREA"}
    return normalize_text(fname) in blocked


def find_field(table, candidates):
    """
    Busca un campo por nombre lógico, ignorando mayúsculas/minúsculas/tildes.
    candidates: lista o set de nombres posibles normalizados.
    """
    lookup = field_names_upper_map(table)
    for c in candidates:
        nc = normalize_text(c)
        if nc in lookup:
            return lookup[nc]
    return None


def find_estado_field(table):
    return find_field(table, {"ESTADO", "STATUS", "ESTATUS"})


def find_expediente_field(table):
    return find_field(table, {"EXPEDIENTE"})


def ensure_field(table, name, ftype, length=None):
    if name in field_names(table):
        return
    if ftype == "TEXT":
        arcpy.management.AddField(table, name, "TEXT", field_length=int(length or 255))
    else:
        arcpy.management.AddField(table, name, ftype)


def ensure_sync_fields(table):
    for fname, ftype, flen in SYNC_FIELDS:
        ensure_field(table, fname, ftype, flen)


def excel_to_stage():
    if exists(STAGE_TABLE):
        arcpy.management.Delete(STAGE_TABLE)

    last_err = None
    for sh in EXCEL_SHEET_CANDIDATES:
        try:
            arcpy.AddMessage(f"Convirtiendo Excel a staging usando hoja: {sh}")
            arcpy.conversion.ExcelToTable(EXCEL_PATH, STAGE_TABLE, sh)
            return sh
        except Exception as e:
            last_err = e

    raise Exception(
        f"No pude leer la hoja del Excel. Probé {EXCEL_SHEET_CANDIDATES}. Error: {last_err}"
    )


def calc_key_in_table(table, expediente_field):
    ensure_field(table, FIELD_KEY, "TEXT", 100)
    with arcpy.da.UpdateCursor(table, [expediente_field, FIELD_KEY]) as cur:
        for exp, _k in cur:
            cur.updateRow((exp, normalize_key(exp)))


def add_stage_fields_to_target(target, stage, exclude_fields=None):
    """
    Agrega al target los campos del Excel/staging que no existan.
    Compara nombres normalizados para no duplicar campos como:
    Expediente / EXPEDIENTE / expediente.
    """
    exclude_fields = set(exclude_fields or [])

    # Mapa de campos existentes en target por nombre normalizado
    tgt_names_norm = {}
    for f in arcpy.ListFields(target):
        tgt_names_norm[normalize_text(f.name)] = f.name

    for f in arcpy.ListFields(stage):
        # saltar campos bloqueados
        if is_blocked_field(f.name):
            continue

        # saltar campos excluidos explícitamente
        if f.name in exclude_fields:
            continue

        # si ya existe un equivalente lógico en target, no agregar otro
        if normalize_text(f.name) in tgt_names_norm:
            continue

        # crear el campo según tipo
        if f.type == "String":
            arcpy.management.AddField(
                target, f.name, "TEXT", field_length=max(10, f.length or 255)
            )
        elif f.type == "Integer":
            arcpy.management.AddField(target, f.name, "LONG")
        elif f.type == "SmallInteger":
            arcpy.management.AddField(target, f.name, "SHORT")
        elif f.type == "Double":
            arcpy.management.AddField(target, f.name, "DOUBLE")
        elif f.type == "Single":
            arcpy.management.AddField(target, f.name, "FLOAT")
        elif f.type == "Date":
            arcpy.management.AddField(target, f.name, "DATE")
        else:
            arcpy.management.AddField(target, f.name, "TEXT", field_length=255)

        arcpy.AddMessage(
            f"Campo nuevo agregado en {os.path.basename(target)}: {f.name} ({f.type})"
        )

        # actualizar mapa para evitar duplicados dentro de la misma corrida
        tgt_names_norm[normalize_text(f.name)] = f.name


def create_attr_table_from_stage():
    arcpy.management.CreateTable(GDB_PATH, TARGET_ATTR_NAME)
    add_stage_fields_to_target(TARGET_ATTR, STAGE_TABLE, exclude_fields={FIELD_KEY})
    ensure_sync_fields(TARGET_ATTR)


def create_index_if_possible(table, field_name, index_name):
    try:
        arcpy.management.AddIndex(
            table, field_name, index_name, "NON_UNIQUE", "ASCENDING"
        )
    except Exception:
        pass


def safe_date_score(v):
    """
    Devuelve un valor comparable para deduplicación.
    Si no es comparable, devuelve None.
    """
    if v is None:
        return None
    return v


def pick_best_date_field(table):
    flds = field_names(table)
    for f in PREFERRED_DATE_FIELDS:
        if f in flds:
            return f
    return None


def deduplicate_target_by_key(table, key_field, log_rows, source_name):
    """
    Elimina duplicados del target por key_field.
    Conserva el registro con fecha preferida más reciente;
    si no aplica, conserva menor OID.
    """
    if key_field not in field_names(table):
        return 0

    best_date_field = pick_best_date_field(table)
    read_fields = ["OID@", key_field]
    if best_date_field:
        read_fields.append(best_date_field)

    winners = {}
    to_delete = set()

    with arcpy.da.SearchCursor(table, read_fields) as cur:
        for row in cur:
            oid = row[0]
            key = row[1]
            if not key:
                continue
            dt_val = row[2] if best_date_field else None

            if key not in winners:
                winners[key] = (oid, dt_val)
                continue

            keep_oid, keep_dt = winners[key]

            score_keep = safe_date_score(keep_dt)
            score_new = safe_date_score(dt_val)

            replace = False
            if score_keep is None and score_new is not None:
                replace = True
            elif score_keep is not None and score_new is None:
                replace = False
            elif score_keep is None and score_new is None:
                replace = oid < keep_oid
            else:
                if score_new > score_keep:
                    replace = True
                elif score_new == score_keep:
                    replace = oid < keep_oid

            if replace:
                to_delete.add(keep_oid)
                winners[key] = (oid, dt_val)
            else:
                to_delete.add(oid)

    if not to_delete:
        return 0

    deleted = 0
    with arcpy.da.UpdateCursor(table, ["OID@", key_field]) as cur:
        for oid, key in cur:
            if oid in to_delete:
                cur.deleteRow()
                deleted += 1
                log_rows.append(
                    [
                        datetime.now().isoformat(timespec="seconds"),
                        "DELETE_DUP",
                        source_name,
                        str(key),
                        f"OID={oid} duplicado eliminado",
                    ]
                )
    return deleted


def collect_source_duplicates(stage_table):
    c = Counter()
    with arcpy.da.SearchCursor(stage_table, [FIELD_KEY, FIELD_EXPEDIENTE_EXCEL]) as cur:
        for key, exp in cur:
            nk = normalize_key(key or exp)
            if nk:
                c[nk] += 1
    return {k: v for k, v in c.items() if v > 1}


def collect_null_empty_estado(stage_table, estado_field):
    rows = []
    if not estado_field:
        return rows
    with arcpy.da.SearchCursor(
        stage_table, [FIELD_EXPEDIENTE_EXCEL, FIELD_KEY, estado_field]
    ) as cur:
        for exp, key, est in cur:
            if est is None or str(est).strip() == "":
                rows.append([exp, key, est])
    return rows


def build_stage_rows_by_key(stage_table, read_fields):
    """
    Si hay duplicados, el último registro del Excel gana para el UPSERT.
    """
    rows_by_key = {}
    skip_rows = []

    with arcpy.da.SearchCursor(stage_table, read_fields) as cur:
        for row in cur:
            data = dict(zip(read_fields, row))
            key = data.get(FIELD_KEY)
            if not key:
                skip_rows.append(data)
                continue
            rows_by_key[str(key)] = data

    return rows_by_key, skip_rows


def reset_report_flags(table):
    flds = field_names(table)
    needed = {"EN_REPORTE", "FECHA_REPORTE"}
    if not needed.issubset(set(flds)):
        return
    with arcpy.da.UpdateCursor(table, ["EN_REPORTE"]) as cur:
        for _ in cur:
            cur.updateRow((0,))


def target_key_to_oid(table):
    if FIELD_KEY not in field_names(table):
        return {}
    d = {}
    with arcpy.da.SearchCursor(table, ["OID@", FIELD_KEY]) as cur:
        for oid, key in cur:
            if key:
                d[str(key)] = oid
    return d


def target_keys_set(table):
    if FIELD_KEY not in field_names(table):
        return set()
    keys = set()
    with arcpy.da.SearchCursor(table, [FIELD_KEY]) as cur:
        for (key,) in cur:
            if key:
                keys.add(str(key))
    return keys


def sql_escape_text(value):
    if value is None:
        return ""
    return str(value).replace("'", "''")


def refresh_output_layers():
    """
    Intenta refrescar las vistas que contengan CM_BASE y CM_ATRIBUTOS.
    No debe interrumpir el ETL si la capa no está presente en el mapa.
    """
    try:
        arcpy.management.ClearWorkspaceCache()
    except Exception:
        pass

    try:
        arcpy.RefreshLayer([TARGET_BASE_NAME, TARGET_ATTR_NAME])
    except Exception:
        try:
            arcpy.RefreshLayer(TARGET_BASE_NAME)
        except Exception:
            pass
        try:
            arcpy.RefreshLayer(TARGET_ATTR_NAME)
        except Exception:
            pass


def transfer_fields_from_stage_to_target(stage_table, target_table):
    """
    Devuelve campos comunes Excel->Target, excluyendo campos internos que se manejan aparte.
    """
    stage_fields = set(field_names(stage_table))
    target_fields = field_names(target_table)

    excluded = {FIELD_KEY, "Fecha_Actualizacion", "EN_REPORTE", "FECHA_REPORTE"}
    result = []
    for f in target_fields:
        if is_blocked_field(f):
            continue
        if f in excluded:
            continue
        if f in stage_fields:
            result.append(f)
    return result


def build_state_lookup_by_key(table):
    estado_field = find_estado_field(table)
    if not estado_field or FIELD_KEY not in field_names(table):
        return {}
    d = {}
    with arcpy.da.SearchCursor(table, [FIELD_KEY, estado_field]) as cur:
        for key, est in cur:
            if key:
                d[str(key)] = normalize_text(est)
    return d


def workbook_write_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title=title[:31])
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)


def save_report_xlsx(
    summary_rows,
    problemas_rows,
    dup_rows,
    null_estado_rows,
    archivado_rows,
    base_sin_excel_rows,
    excel_sin_base_rows,
    log_attr_rows,
    log_base_rows,
    out_path,
):
    wb = Workbook()
    wb.remove(wb.active)

    workbook_write_sheet(wb, "RESUMEN", ["METRICA", "VALOR"], summary_rows)
    workbook_write_sheet(
        wb,
        "PROBLEMAS_EXPEDIENTE",
        ["EXPEDIENTE", "EXPEDIENTE_KEY", "TIPO_PROBLEMA", "DETALLE"],
        problemas_rows,
    )
    workbook_write_sheet(
        wb, "DUPLICADOS_ORIGEN", ["EXPEDIENTE_KEY", "CANTIDAD"], dup_rows
    )
    workbook_write_sheet(
        wb,
        "ESTADOS_VACIOS",
        ["EXPEDIENTE", "EXPEDIENTE_KEY", "ESTADO"],
        null_estado_rows,
    )
    workbook_write_sheet(
        wb,
        "NUEVOS_ARCHIVADO",
        ["EXPEDIENTE", "EXPEDIENTE_KEY", "ESTADO_GIS_ACTUAL", "ESTADO_EXCEL_NUEVO"],
        archivado_rows,
    )
    workbook_write_sheet(
        wb,
        "CM_BASE_SIN_EXCEL",
        ["EXPEDIENTE_CM_BASE", "EXPEDIENTE_KEY"],
        base_sin_excel_rows,
    )
    workbook_write_sheet(
        wb,
        "EXCEL_SIN_CM_BASE",
        ["EXPEDIENTE_EXCEL", "EXPEDIENTE_KEY", "ESTADO_EXCEL"],
        excel_sin_base_rows,
    )
    workbook_write_sheet(
        wb,
        "LOG_CM_ATRIBUTOS",
        ["TIMESTAMP", "ACCION", "TABLA", "EXPEDIENTE_KEY", "DETALLE"],
        log_attr_rows,
    )
    workbook_write_sheet(
        wb,
        "LOG_CM_BASE",
        ["TIMESTAMP", "ACCION", "TABLA", "EXPEDIENTE_KEY", "DETALLE"],
        log_base_rows,
    )

    wb.save(out_path)


# =========================================================
# EJECUCIÓN PRINCIPAL
# =========================================================
try:
    if not os.path.exists(EXCEL_PATH):
        raise Exception(f"No existe el Excel en: {EXCEL_PATH}")
    if not exists(TARGET_BASE):
        raise Exception(f"No existe {TARGET_BASE_NAME} en la GDB actual.")
    if not arcpy.env.scratchGDB or not exists(arcpy.env.scratchGDB):
        raise Exception("No existe scratchGDB disponible para el staging.")

    # estructuras para reporte
    log_attr = []
    log_base = []
    problemas = []

    # 1) Excel -> staging
    sheet_used = excel_to_stage()

    # 2) validar campo expediente en excel
    expediente_stage = find_expediente_field(STAGE_TABLE)
    if not expediente_stage:
        raise Exception(
            f"El Excel no trae el campo obligatorio '{FIELD_EXPEDIENTE_EXCEL}'."
        )

    # 3) calcular key en staging
    calc_key_in_table(STAGE_TABLE, expediente_stage)

    # 4) controles de calidad del origen
    estado_stage = find_estado_field(STAGE_TABLE)
    dup_dict = collect_source_duplicates(STAGE_TABLE)
    dup_rows = sorted(
        [[k, v] for k, v in dup_dict.items()], key=lambda x: (-x[1], x[0])
    )

    null_estado_rows = collect_null_empty_estado(STAGE_TABLE, estado_stage)

    for k, v in dup_rows:
        problemas.append(["", k, "DUPLICADO_ORIGEN", f"Cantidad={v}"])
    for exp, key, est in null_estado_rows:
        problemas.append([exp, key, "ESTADO_VACIO", "Estado nulo o vacío en Excel"])

    # 5) crear CM_ATRIBUTOS si no existe
    if not exists(TARGET_ATTR):
        arcpy.AddMessage(f"{TARGET_ATTR_NAME} no existe. Creándola...")
        create_attr_table_from_stage()
        create_index_if_possible(TARGET_ATTR, FIELD_KEY, "IX_ATTR_EXP_KEY")

    # 6) asegurar campos en targets
    add_stage_fields_to_target(TARGET_ATTR, STAGE_TABLE, exclude_fields={FIELD_KEY})
    ensure_sync_fields(TARGET_ATTR)

    add_stage_fields_to_target(TARGET_BASE, STAGE_TABLE, exclude_fields={FIELD_KEY})
    ensure_sync_fields(TARGET_BASE)

    # 7) calcular key en CM_BASE si no la tiene o si hay valores vacíos
    expediente_base = find_expediente_field(TARGET_BASE)
    if not expediente_base:
        raise Exception(
            f"{TARGET_BASE_NAME} no tiene un campo de expediente reconocible."
        )

    with arcpy.da.UpdateCursor(TARGET_BASE, [expediente_base, FIELD_KEY]) as cur:
        for exp, key in cur:
            nk = normalize_key(exp)
            if key != nk:
                cur.updateRow((exp, nk))

    # 8) dedupe residuos en targets
    deleted_attr_dups = deduplicate_target_by_key(
        TARGET_ATTR, FIELD_KEY, log_attr, TARGET_ATTR_NAME
    )
    deleted_base_dups = deduplicate_target_by_key(
        TARGET_BASE, FIELD_KEY, log_base, TARGET_BASE_NAME
    )

    # 9) reset flags
    reset_report_flags(TARGET_ATTR)
    reset_report_flags(TARGET_BASE)

    # 10) staging en memoria
    stage_read_fields = [FIELD_KEY] + [
        f.name
        for f in arcpy.ListFields(STAGE_TABLE)
        if not is_blocked_field(f.name) and f.name != FIELD_KEY
    ]
    stage_rows_by_key, skipped_stage_rows = build_stage_rows_by_key(
        STAGE_TABLE, stage_read_fields
    )

    for data in skipped_stage_rows:
        problemas.append(
            [
                data.get(expediente_stage, ""),
                "",
                "LLAVE_VACIA",
                "No se pudo construir EXPEDIENTE_KEY en staging",
            ]
        )

    # 11) lookups de targets
    attr_key_oid = target_key_to_oid(TARGET_ATTR)
    base_keys = target_keys_set(TARGET_BASE)

    # 12) detectar nuevos ARCHIVADO SOLO si el expediente existe en CM_BASE
    estado_base = find_estado_field(TARGET_BASE)
    estado_actual_base = build_state_lookup_by_key(TARGET_BASE)
    archivado_rows = []

    for key, data in stage_rows_by_key.items():
        # Si no existe en CM_BASE, NO puede considerarse nuevo archivado en GIS
        if key not in base_keys:
            continue

        estado_nuevo = normalize_text(data.get(estado_stage))
        estado_viejo = normalize_text(estado_actual_base.get(key, ""))

        if estado_nuevo == VALOR_ARCHIVADO and estado_viejo != VALOR_ARCHIVADO:
            exp = data.get(expediente_stage)
            archivado_rows.append([exp, key, estado_viejo, estado_nuevo])
            problemas.append(
                [
                    exp,
                    key,
                    "NUEVO_ARCHIVADO",
                    f"Antes='{estado_viejo}' Ahora='{estado_nuevo}'",
                ]
            )

    # 13) upsert CM_ATRIBUTOS
    transfer_attr_fields = transfer_fields_from_stage_to_target(
        STAGE_TABLE, TARGET_ATTR
    )
    now_dt = datetime.now()
    attr_updates = 0
    attr_inserts = 0
    attr_nochange = 0

    for key, data in stage_rows_by_key.items():
        if key in attr_key_oid:
            oid = attr_key_oid[key]
            upd_fields = (
                ["OID@"]
                + transfer_attr_fields
                + ["Fecha_Actualizacion", "EN_REPORTE", "FECHA_REPORTE"]
            )
            where = f"OBJECTID = {oid}"

            with arcpy.da.UpdateCursor(
                TARGET_ATTR, upd_fields, where_clause=where
            ) as cur:
                for row in cur:
                    changed = False
                    new_vals = [row[0]]

                    for i, fld in enumerate(transfer_attr_fields, start=1):
                        new_val = data.get(fld)
                        old_val = row[i]
                        if new_val != old_val:
                            changed = True
                        new_vals.append(new_val)

                    new_vals.append(
                        now_dt if changed else row[len(transfer_attr_fields) + 1]
                    )
                    new_vals.append(1)
                    new_vals.append(now_dt)

                    cur.updateRow(new_vals)

                    if changed:
                        attr_updates += 1
                        log_attr.append(
                            [
                                datetime.now().isoformat(timespec="seconds"),
                                "UPDATE",
                                TARGET_ATTR_NAME,
                                key,
                                "Registro actualizado",
                            ]
                        )
                    else:
                        attr_nochange += 1
                        log_attr.append(
                            [
                                datetime.now().isoformat(timespec="seconds"),
                                "NOCHANGE",
                                TARGET_ATTR_NAME,
                                key,
                                "Sin cambios de atributos",
                            ]
                        )
        else:
            ins_fields = transfer_attr_fields + [
                "Fecha_Actualizacion",
                "EN_REPORTE",
                "FECHA_REPORTE",
            ]
            ins_vals = [data.get(f) for f in transfer_attr_fields] + [now_dt, 1, now_dt]

            with arcpy.da.InsertCursor(TARGET_ATTR, ins_fields) as cur:
                cur.insertRow(ins_vals)

            attr_inserts += 1
            log_attr.append(
                [
                    datetime.now().isoformat(timespec="seconds"),
                    "INSERT",
                    TARGET_ATTR_NAME,
                    key,
                    "Registro insertado",
                ]
            )

    # 14) upsert CM_BASE (real, con todos los campos del Excel)
    transfer_base_fields = transfer_fields_from_stage_to_target(
        STAGE_TABLE, TARGET_BASE
    )
    base_updates = 0
    base_inserts = 0
    base_nochange = 0
    base_missing_in_excel = []
    excel_missing_in_base = []

    # Como CM_BASE es capa espacial, no insertamos nuevos registros si no hay geometría.
    # Pero sí actualizamos TODOS los registros que tengan el mismo EXPEDIENTE_KEY.
    key_field_delimited = arcpy.AddFieldDelimiters(TARGET_BASE, FIELD_KEY)

    for key, data in stage_rows_by_key.items():
        if key in base_keys:
            upd_fields = transfer_base_fields + [
                "Fecha_Actualizacion",
                "EN_REPORTE",
                "FECHA_REPORTE",
            ]
            where = f"{key_field_delimited} = '{sql_escape_text(key)}'"

            updated_any = False

            with arcpy.da.UpdateCursor(
                TARGET_BASE, upd_fields, where_clause=where
            ) as cur:
                for row in cur:
                    changed = False
                    new_vals = []

                    for i, fld in enumerate(transfer_base_fields):
                        new_val = data.get(fld)
                        old_val = row[i]
                        if new_val != old_val:
                            changed = True
                        new_vals.append(new_val)

                    old_fecha = row[len(transfer_base_fields)]
                    new_vals.append(now_dt if changed else old_fecha)
                    new_vals.append(1)
                    new_vals.append(now_dt)

                    cur.updateRow(new_vals)
                    updated_any = True

                    if changed:
                        base_updates += 1
                        log_base.append(
                            [
                                datetime.now().isoformat(timespec="seconds"),
                                "UPDATE",
                                TARGET_BASE_NAME,
                                key,
                                "Registro espacial actualizado",
                            ]
                        )
                    else:
                        base_nochange += 1
                        log_base.append(
                            [
                                datetime.now().isoformat(timespec="seconds"),
                                "NOCHANGE",
                                TARGET_BASE_NAME,
                                key,
                                "Sin cambios de atributos",
                            ]
                        )

            if not updated_any:
                estado_excel = data.get(estado_stage) if estado_stage else None
                excel_missing_in_base.append(
                    [data.get(expediente_stage), key, estado_excel]
                )
                problemas.append(
                    [
                        data.get(expediente_stage),
                        key,
                        "EXCEL_SIN_CM_BASE",
                        f"Existe en Excel pero no se pudo actualizar en CM_BASE. Estado Excel='{estado_excel}'",
                    ]
                )
                log_base.append(
                    [
                        datetime.now().isoformat(timespec="seconds"),
                        "SKIP_NO_GEOMETRY",
                        TARGET_BASE_NAME,
                        key,
                        f"Existe llave en CM_BASE pero no se pudo actualizar. Estado Excel='{estado_excel}'",
                    ]
                )
        else:
            estado_excel = data.get(estado_stage) if estado_stage else None
            excel_missing_in_base.append(
                [data.get(expediente_stage), key, estado_excel]
            )
            problemas.append(
                [
                    data.get(expediente_stage),
                    key,
                    "EXCEL_SIN_CM_BASE",
                    f"Existe en Excel pero no en CM_BASE. Estado Excel='{estado_excel}'",
                ]
            )
            log_base.append(
                [
                    datetime.now().isoformat(timespec="seconds"),
                    "SKIP_NO_GEOMETRY",
                    TARGET_BASE_NAME,
                    key,
                    f"No existe registro espacial en CM_BASE. Estado Excel='{estado_excel}'",
                ]
            )

    # 15) marcar SIN_DATOS en CM_ATRIBUTOS y CM_BASE donde no vinieron en Excel
    attr_sin_datos = 0
    base_sin_datos = 0

    estado_attr = find_estado_field(TARGET_ATTR)
    if estado_attr:
        with arcpy.da.UpdateCursor(
            TARGET_ATTR, ["EN_REPORTE", estado_attr, FIELD_KEY]
        ) as cur:
            for en_rep, est, key in cur:
                if en_rep == 0:
                    if est is None or str(est).strip() != VALOR_SIN_DATOS:
                        cur.updateRow((en_rep, VALOR_SIN_DATOS, key))
                        attr_sin_datos += 1
                        log_attr.append(
                            [
                                datetime.now().isoformat(timespec="seconds"),
                                "SET_SIN_DATOS",
                                TARGET_ATTR_NAME,
                                str(key),
                                "No vino en Excel",
                            ]
                        )
                        problemas.append(
                            [
                                "",
                                str(key),
                                "ATRIBUTO_SIN_DATOS",
                                "No vino en Excel actual",
                            ]
                        )

    # Para CM_BASE, determinar "sin Excel" por ausencia REAL de la llave en el Excel,
    # no solo por EN_REPORTE=0.
    excel_keys = set(stage_rows_by_key.keys())
    reported_missing_base_keys = set()

    if estado_base:
        with arcpy.da.UpdateCursor(
            TARGET_BASE, [expediente_base, FIELD_KEY, estado_base]
        ) as cur:
            for exp, key, est in cur:
                key_str = str(key) if key else ""
                if key_str and key_str not in excel_keys:
                    if est is None or str(est).strip() != VALOR_SIN_DATOS:
                        cur.updateRow((exp, key, VALOR_SIN_DATOS))

                    base_sin_datos += 1

                    if key_str not in reported_missing_base_keys:
                        base_missing_in_excel.append([exp, key_str])
                        log_base.append(
                            [
                                datetime.now().isoformat(timespec="seconds"),
                                "SET_SIN_DATOS",
                                TARGET_BASE_NAME,
                                key_str,
                                "Existe en CM_BASE pero no vino en Excel",
                            ]
                        )
                        problemas.append(
                            [
                                exp,
                                key_str,
                                "CM_BASE_SIN_EXCEL",
                                "Existe en CM_BASE pero no vino en Excel actual",
                            ]
                        )
                        reported_missing_base_keys.add(key_str)

    # 16) resumen y reporte único
    summary_rows = [
        ["HOJA_USADA", sheet_used],
        ["EXCEL_PATH", EXCEL_PATH],
        ["GDB_PATH", GDB_PATH],
        ["CM_ATRIBUTOS", TARGET_ATTR],
        ["CM_BASE", TARGET_BASE],
        ["DUPLICADOS_ORIGEN", len(dup_rows)],
        ["ESTADOS_VACIOS", len(null_estado_rows)],
        ["NUEVOS_ARCHIVADO", len(archivado_rows)],
        ["CM_BASE_SIN_EXCEL", len(base_missing_in_excel)],
        ["EXCEL_SIN_CM_BASE", len(excel_missing_in_base)],
        ["SKIPS_STAGING_SIN_LLAVE", len(skipped_stage_rows)],
        ["CM_ATRIBUTOS_DUPLICADOS_ELIMINADOS", deleted_attr_dups],
        ["CM_BASE_DUPLICADOS_ELIMINADOS", deleted_base_dups],
        ["CM_ATRIBUTOS_UPDATES", attr_updates],
        ["CM_ATRIBUTOS_INSERTS", attr_inserts],
        ["CM_ATRIBUTOS_NOCHANGE", attr_nochange],
        ["CM_ATRIBUTOS_SET_SIN_DATOS", attr_sin_datos],
        ["CM_BASE_UPDATES", base_updates],
        ["CM_BASE_INSERTS", base_inserts],
        ["CM_BASE_NOCHANGE", base_nochange],
        ["CM_BASE_SET_SIN_DATOS", base_sin_datos],
        ["REPORTE_XLSX", REPORT_XLSX],
    ]

    save_report_xlsx(
        summary_rows=summary_rows,
        problemas_rows=problemas,
        dup_rows=dup_rows,
        null_estado_rows=null_estado_rows,
        archivado_rows=archivado_rows,
        base_sin_excel_rows=base_missing_in_excel,
        excel_sin_base_rows=excel_missing_in_base,
        log_attr_rows=log_attr,
        log_base_rows=log_base,
        out_path=REPORT_XLSX,
    )

    refresh_output_layers()

    arcpy.AddMessage("======================================")
    arcpy.AddMessage("ETL ÚNICO COMPLETADO")
    arcpy.AddMessage("======================================")
    arcpy.AddMessage(f"CM_ATRIBUTOS Updates: {attr_updates}")
    arcpy.AddMessage(f"CM_ATRIBUTOS Inserts: {attr_inserts}")
    arcpy.AddMessage(f"CM_BASE Updates: {base_updates}")
    arcpy.AddMessage(f"CM_BASE Sin datos: {base_sin_datos}")
    arcpy.AddMessage(f"Reporte Excel: {REPORT_XLSX}")

except Exception as e:
    import traceback

    arcpy.AddError("ERROR EN ETL ÚNICO")
    arcpy.AddError(str(e))
    arcpy.AddError(traceback.format_exc())
    raise
